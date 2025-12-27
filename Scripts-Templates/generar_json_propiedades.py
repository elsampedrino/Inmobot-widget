"""
Script para procesar el Excel de BBR Grupo Inmobiliario
y generar JSON de propiedades en el formato del bot

Autor: Claude Code
Fecha: 2025-01-13
"""

import openpyxl
import json
import os
import re
from pathlib import Path

# ============================================
# CONFIGURACIÓN
# ============================================

EXCEL_PATH = "BASE DE DATOS PROPIEDADES BOOT.xlsx"
OUTPUT_JSON = "propiedades_bbr.json"
FOTOS_BASE_DIR = "PROPIEDADES"

# Mapeo de hojas del Excel a tipo de propiedad
HOJA_TO_TIPO = {
    "Casas": "Casa",
    "Deptos": "Departamento",
    "Lotes-Terrenos": "Terreno",
    "Campos": "Campo",
    "Alquiler": None  # Se determina por la columna "Tipo"
}

# ============================================
# FUNCIONES AUXILIARES
# ============================================

def normalizar_nombre_carpeta(direccion):
    """
    Normaliza el nombre de dirección para buscar carpeta de fotos
    Ej: "Colón al 1100" -> "Colon al 1100" o "Colon 1100"
    """
    if not direccion:
        return None

    # Quitar tildes y caracteres especiales
    reemplazos = {
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u',
        'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U',
        'ñ': 'n', 'Ñ': 'N'
    }

    direccion_normalizada = direccion
    for viejo, nuevo in reemplazos.items():
        direccion_normalizada = direccion_normalizada.replace(viejo, nuevo)

    return direccion_normalizada.strip()

def buscar_carpeta_fotos(direccion, tipo_propiedad):
    """
    Busca la carpeta de fotos correspondiente a una propiedad
    """
    if not direccion:
        return None

    # Determinar carpeta base según tipo
    carpeta_tipo_map = {
        "Casa": "Casas",
        "Departamento": "Alquiler",  # Los deptos están en Alquiler
        "Terreno": "Lotes",
        "Campo": "Campos",
        "Local": "Alquiler"
    }

    carpeta_tipo = carpeta_tipo_map.get(tipo_propiedad, "Casas")
    base_path = Path(FOTOS_BASE_DIR) / carpeta_tipo

    if not base_path.exists():
        return None

    # Normalizar dirección
    direccion_norm = normalizar_nombre_carpeta(direccion)

    # Buscar carpeta que coincida
    for carpeta in base_path.iterdir():
        if carpeta.is_dir():
            carpeta_norm = normalizar_nombre_carpeta(carpeta.name)

            # Comparación flexible
            if (direccion_norm.lower() in carpeta_norm.lower() or
                carpeta_norm.lower() in direccion_norm.lower()):
                return carpeta

    return None

def obtener_fotos_de_carpeta(carpeta_path):
    """
    Obtiene lista de archivos de fotos en una carpeta
    """
    if not carpeta_path or not carpeta_path.exists():
        return []

    extensiones_validas = {'.jpg', '.jpeg', '.png', '.webp'}
    fotos = []

    for archivo in sorted(carpeta_path.iterdir()):
        if archivo.suffix.lower() in extensiones_validas:
            # Ruta relativa desde PROPIEDADES
            ruta_relativa = archivo.relative_to(Path(FOTOS_BASE_DIR))
            fotos.append(str(ruta_relativa).replace('\\', '/'))

    return fotos

def limpiar_valor_precio(valor_str):
    """
    Limpia y extrae el valor numérico del precio
    Ej: "u$S 139.000" -> 139000
    Ej: "u$S 140.000 - 11 x 42" -> 140000
    """
    if not valor_str or valor_str == "":
        return None

    # Convertir a string si es número
    valor_str = str(valor_str)

    # Si tiene guión o "x" (dimensiones), tomar solo la parte antes del guión
    if '-' in valor_str:
        valor_str = valor_str.split('-')[0].strip()

    # Buscar patrón de precio (números con puntos o comas como separadores)
    # Ej: "139.000", "139,000", "139000"
    match = re.search(r'(\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{1,2})?)', valor_str)

    if match:
        # Extraer el número y quitar puntos/comas
        precio = match.group(1)
        precio = precio.replace('.', '').replace(',', '')
        return int(precio)

    return None

def extraer_superficie(valor):
    """
    Extrae superficie en m2
    """
    if not valor or valor == "":
        return None

    try:
        return int(float(valor))
    except:
        # Intentar extraer números
        numeros = re.findall(r'\d+', str(valor))
        if numeros:
            return int(numeros[0])
        return None

def generar_id_propiedad(index, tipo):
    """
    Genera ID único para propiedad
    Ej: BBR-CASA-001, BBR-DEPTO-015
    """
    tipo_codigo = tipo[:4].upper() if tipo else "PROP"
    return f"BBR-{tipo_codigo}-{str(index + 1).zfill(3)}"

# ============================================
# PROCESAMIENTO PRINCIPAL
# ============================================

def procesar_excel():
    """
    Lee el Excel y genera JSON de propiedades
    """
    print("📄 Leyendo Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    propiedades = []
    index_global = 0

    # Procesar cada hoja
    for nombre_hoja in wb.sheetnames:
        if nombre_hoja not in HOJA_TO_TIPO:
            print(f"⏭️  Saltando hoja: {nombre_hoja}")
            continue

        print(f"\n📋 Procesando hoja: {nombre_hoja}")
        sheet = wb[nombre_hoja]

        # Leer filas (saltar header en fila 1)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Extraer valores de columnas CORREGIDAS
            num = row[0] if len(row) > 0 else None
            operacion = row[1] if len(row) > 1 else None
            tipo = row[2] if len(row) > 2 else HOJA_TO_TIPO.get(nombre_hoja)
            # Columna 3 tiene imagen, la saltamos
            direccion = row[4] if len(row) > 4 else None
            localidad = row[5] if len(row) > 5 else None
            superficie_construida = row[6] if len(row) > 6 else None  # Construida m2
            lote_terreno = row[7] if len(row) > 7 else None  # Lote/Terreno m2 (ej: "11 x 42")
            descripcion = row[8] if len(row) > 8 else None  # Descripción
            valor_str = row[9] if len(row) > 9 else None  # Valor (precio)
            dueno = row[10] if len(row) > 10 else None

            # Validar que tenga datos mínimos
            if not direccion or not operacion:
                continue

            # Generar ID
            prop_id = generar_id_propiedad(index_global, tipo)
            index_global += 1

            # Buscar carpeta de fotos
            carpeta_fotos = buscar_carpeta_fotos(direccion, tipo)
            fotos_paths = obtener_fotos_de_carpeta(carpeta_fotos) if carpeta_fotos else []

            # Limpiar precio
            valor_limpio = limpiar_valor_precio(valor_str)

            # Extraer superficie construida
            superficie = extraer_superficie(superficie_construida)

            # Construir objeto propiedad
            propiedad = {
                "id": prop_id,
                "tipo": tipo,
                "operacion": operacion,
                "titulo": f"{tipo} en {direccion}" if direccion else f"{tipo}",
                "direccion": {
                    "calle": direccion,
                    "localidad": localidad if localidad else "Ramallo",
                    "provincia": "Buenos Aires",
                    "pais": "Argentina"
                },
                "precio": {
                    "valor": valor_limpio,
                    "moneda": "USD"
                },
                "caracteristicas": {
                    "superficie_total": superficie
                },
                "descripcion": descripcion if descripcion else f"{tipo} en {direccion}",
                "disponibilidad": "Disponible",
                "fotos": {
                    "cantidad": len(fotos_paths),
                    "paths": fotos_paths  # Rutas locales, se convertirán a URLs de Cloudinary
                }
            }

            propiedades.append(propiedad)
            print(f"  ✅ {prop_id}: {direccion} ({len(fotos_paths)} fotos)")

    wb.close()

    # Generar JSON final
    output = {
        "propiedades": propiedades,
        "total": len(propiedades),
        "ultima_actualizacion": "2025-01-13"
    }

    # Guardar JSON
    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\n✅ JSON generado: {OUTPUT_JSON}")
    print(f"📊 Total de propiedades: {len(propiedades)}")

    # Estadísticas
    print("\n📈 Estadísticas:")
    tipos = {}
    for prop in propiedades:
        tipo = prop['tipo']
        tipos[tipo] = tipos.get(tipo, 0) + 1

    for tipo, cantidad in tipos.items():
        print(f"  - {tipo}: {cantidad}")

# ============================================
# EJECUCIÓN
# ============================================

if __name__ == "__main__":
    import sys
    import io

    # Fix encoding for Windows
    if sys.platform == 'win32':
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

    print("🏠 BBR Grupo Inmobiliario - Generador de JSON\n")
    print("=" * 60)

    try:
        procesar_excel()
        print("\n" + "=" * 60)
        print("✅ Proceso completado exitosamente!")
        print("\n📝 Próximos pasos:")
        print("  1. Revisar propiedades_bbr.json")
        print("  2. Ejecutar script de optimización de imágenes")
        print("  3. Subir imágenes a Cloudinary")
        print("  4. Actualizar URLs en JSON con las de Cloudinary")

    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
