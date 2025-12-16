"""
Script para ver la estructura completa de una fila del Excel
"""
import openpyxl
import sys
import io

# Fix encoding for Windows
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

EXCEL_PATH = "BASE DE DATOS PROPIEDADES BOOT.xlsx"

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

print("=" * 100)
print("ESTRUCTURA COMPLETA - HOJA: Casas")
print("=" * 100)

sheet = wb["Casas"]

# Ver headers (fila 1)
print("\n📋 HEADERS (Fila 1):")
print("-" * 100)
headers = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
for idx, header in enumerate(headers):
    if header:
        print(f"Columna {idx}: {header}")

# Ver datos de la primera propiedad (fila 2)
print("\n\n📄 DATOS PRIMERA PROPIEDAD (Fila 2):")
print("-" * 100)
row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
for idx, valor in enumerate(row):
    header_name = headers[idx] if idx < len(headers) else f"Col_{idx}"
    if valor:
        # Truncar valores largos
        valor_str = str(valor)
        if len(valor_str) > 100:
            valor_str = valor_str[:100] + "..."
        print(f"Columna {idx} ({header_name}): {valor_str}")

wb.close()
